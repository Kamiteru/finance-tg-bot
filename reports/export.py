from services.transaction import get_last_transactions
from services.user import get_user_currency, format_amount_with_currency
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfutils
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl
import io
from sqlalchemy.orm import joinedload
import datetime
import os

def export_pdf(user_id: int, limit: int = 100):
    # Export last transactions to PDF for specific user
    try:
        txs = get_last_transactions(user_id, limit)
        if not txs:
            return None
        
        user_currency = get_user_currency(user_id)
        
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        width, height = A4
        
        # Header
        y = height - 40
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, y, "Financial Report")
        
        y -= 25
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
        c.drawString(40, y-15, f"Total transactions: {len(txs)}")
        c.drawString(40, y-30, f"Currency: {user_currency}")
        
        # Table header
        y -= 60
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, "Date")
        c.drawString(120, y, "Category")
        c.drawString(220, y, "Type")
        c.drawString(280, y, f"Amount ({user_currency})")
        
        # Draw line
        y -= 5
        c.line(40, y, 450, y)
        
        # Transactions
        y -= 15
        c.setFont("Helvetica", 9)
        total_income = 0
        total_expense = 0
        
        for t in txs:
            if y < 60:  # New page
                c.showPage()
                y = height - 40
                
            cat = t.category.name if t.category else "No category"
            type_str = "Income" if t.type == "income" else "Expense"
            amount = float(t.amount)
            
            if t.type == "income":
                total_income += amount
            else:
                total_expense += amount
            
            c.drawString(40, y, t.date.strftime('%d.%m.%Y'))
            c.drawString(120, y, cat[:12])
            c.drawString(220, y, type_str)
            c.drawString(280, y, f"{amount:,.2f}")
            y -= 12
        
        # Summary
        y -= 20
        c.line(40, y, 450, y)
        y -= 15
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, f"Total Income: {total_income:,.2f} {user_currency}")
        c.drawString(250, y, f"Total Expense: {total_expense:,.2f} {user_currency}")
        c.drawString(40, y-15, f"Balance: {total_income - total_expense:,.2f} {user_currency}")
        
        c.save()
        buf.seek(0)
        return buf
    except Exception as e:
        print(f"Error generating PDF: {e}")
        return None

def export_excel(user_id: int, limit: int = 100):
    # Export last transactions to Excel for specific user
    try:
        txs = get_last_transactions(user_id, limit)
        if not txs:
            return None
        
        user_currency = get_user_currency(user_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Headers
        headers = ["Date", "Category", "Type", f"Amount ({user_currency})", "Description"]
        ws.append(headers)
        
        # Style header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        total_income = 0
        total_expense = 0
        
        for t in txs:
            cat = t.category.name if t.category else "No category"
            type_str = "Income" if t.type == "income" else "Expense"
            amount = float(t.amount)
            
            if t.type == "income":
                total_income += amount
            else:
                total_expense += amount
            
            ws.append([
                t.date.strftime('%d.%m.%Y %H:%M'),
                cat,
                type_str,
                amount,
                t.description or ""
            ])
        
        # Summary
        ws.append([])
        ws.append(["TOTAL:", "", "", "", ""])
        ws.append(["Income:", "", "", total_income, ""])
        ws.append(["Expenses:", "", "", total_expense, ""])
        ws.append(["Balance:", "", "", total_income - total_expense, ""])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    except Exception as e:
        print(f"Error generating Excel: {e}")
        return None 